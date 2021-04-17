import openpyxl
county_schema_path = "../data/AHRF_2019-2020/DOC/AHRF 2019-2020 Technical Documentation.xlsx"
def parse_schema(schema_path, mode):
    if mode == 'state':
        schema_file = openpyxl.load_workbook(schema_path)
        sheet = schema_file.active

        data_values = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if (row[0]):
                if (row[0].startswith('SF')):
                    data = {}
                    data['field'] = row[0].strip()
                    index_values = row[1].split(' - ')
                    data['start'] = int(index_values[0])
                    data['end'] = int(index_values[1])
                    data['year'] = row[2]
                    data['variable_name'] = row[3]
                    data['characteristics'] = row[4]
                    data['source'] = row[5]
                    data['dateon'] = row[6].strip()
                    data_values.append(data)

    elif mode == 'county':
        schema_file = openpyxl.load_workbook(schema_path)
        sheet = schema_file.active

        data_values = []

        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):

            if (row[0]):
                try:
                    if (row[0].startswith(' F')):
                        data = {}
                        data['field'] = row[0].strip()
                        index_values = row[1].split('-')
                        data['start'] = int(index_values[0])
                        data['end'] = int(index_values[1])
                        data['year'] = row[2]
                        data['variable_name'] = row[3]
                        data['characteristics'] = row[4]
                        data['source'] = row[5]
                        data['dateon'] = row[6]
                        data_values.append(data)
                except:
                    pass
                    
                    
    else:
        raise Exception(f"{mode} is not valid. Valid modes - county or state")


    return data_values


county_schema = parse_schema(county_schema_path, 'county')
print (county_schema[:20])